import time
from asyncio import wait
from tokenize import tabsize

import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.devtools.v141.dom_storage import Item
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

chrome_options = Options()
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--ignore-certificate-errors")
# chrome_options.add_argument("headless")63.Handle_Excel_File.py
# chrome_options.add_argument("--window-size=1920,1080")
# chrome_options.add_argument("--disable-notifications")
# chrome_options.add_argument("--no-sandbox")
# chrome_options.add_argument("--disable-gpu")
# chrome_options.add_argument("--lang=vi")
# chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
# chrome_options.add_experimental_option("useAutomationExtension", False)

def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(filePath)

service = Service("D:/OneDrive/Learning/Python_Selenium/chromedriver-win64/chromedriver.exe")
try:
    driver = webdriver.Chrome(
        service=service,
        options=chrome_options
    )
    driver.implicitly_wait(3)

    excel_file_obj = openpyxl.load_workbook("data.xlsx")
    # username = excel_file_obj["Sheet1"]["A2"].value
    # password = excel_file_obj["Sheet1"]["B2"].value

    driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
    file_path = "C:\\Users\\Phu\\Desktop\\Py_Selenium\\download.xlsx"
    driver.find_element(By.ID, "downloadButton").click()

    fruit_name = "Apple"
    # Update excel
    update_excel_data(file_path, "Apple", "price", 9999)

    # Upload excel
    driver.find_element(By.CSS_SELECTOR, "input[type='file']").send_keys(file_path)
    WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")))
    priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
    actual_price = driver.find_element(By.XPATH,"//div[text()='" + fruit_name + "']/parent::div/parent::div/div[@id='cell-" + priceColumn + "-undefined']").text
    assert "9999" == actual_price


finally:
    driver.quit()